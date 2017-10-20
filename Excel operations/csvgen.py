import psycopg2
import csv
#from users.models import Order
#from rest_framework import viewsets
#from rest_framework.views import APIView
#import views
import openpyxl
import datetime
from openpyxl import Workbook
from openpyxl import load_workbook
import os.path
import os

def csvg(details):
	d = datetime.datetime.now()
	date = str(d.day) + "-" + str(d.month) + "-" + str(d.year)
	stud_id = details[0]
	cost = details[1]

	try:
		wb = load_workbook(filename = '/home/ravi/'+date+'.xlsx')
		ws = wb.active
		cola = ws['A']
		print stud_id, cost
		flag = 1
		for i in cola:
		         if i.value == stud_id:
				flag = 0
				row = str(i.row)
				column = i.column
				colb = ws['B']
				for j in colb:
					if j.column+str(j.row) == 'B'+row:
						j.value = int(j.value) + int(cost)
						wb.save(filename = '/home/ravi/'+date+'.xlsx')
						break

		if flag == 1:
			ws.append([stud_id, cost])
			wb.save(filename = '/home/ravi/'+date+'.xlsx')
		wb.save(filename = '/home/ravi/'+date+'.xlsx')
	except Exception as e:
#		print e
		print "Workbook generated at "+ '/home/ravi/'+date+'.xlsx'
		wb = Workbook()
		ws = wb.active
		ws.append([stud_id,cost])
		wb.save(filename = '/home/ravi/'+date+'.xlsx')
